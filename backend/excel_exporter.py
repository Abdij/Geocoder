from __future__ import annotations

import importlib.util
from pathlib import Path
from typing import Iterable

import pandas as pd

from backend.district_summary import build_summary_tables, district_groups
from backend.utils import detect_column_map, output_path, truncate_sheet_name
from config import CLUSTER_COLORS

# Longest key first so "child protection" matches before the bare
# "protection" substring it also contains.
_CLUSTER_COLOR_KEYS = sorted(CLUSTER_COLORS, key=len, reverse=True)


def _cluster_color(cluster_name: str) -> str | None:
    """Resolve a real-world cluster label to a configured color by substring.

    Response data rarely uses the bare cluster name from CLUSTER_COLORS
    ("Food Security Cluster", "Shelter and NFIs", "Gender Based Violence"),
    so an exact dict lookup misses most real values. This checks whether any
    known cluster key appears anywhere in the (lowercased) label instead.
    """
    normalized = cluster_name.strip().lower()
    for key in _CLUSTER_COLOR_KEYS:
        if key in normalized:
            return CLUSTER_COLORS[key]
    return None


HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
LIGHT_FILL = "D9EAF7"
WIDTH_SCAN_ROW_LIMIT = 250
EXCEL_OUTPUT_KEYS = {"Cleaned Excel", "District Workbook", "District Summary"}


def _validation_frames(validation_report: dict[str, object] | None) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not validation_report:
        return pd.DataFrame(), pd.DataFrame()
    metrics = validation_report.get("metrics", {})
    metric_df = pd.DataFrame(
        [{"Metric": key.replace("_", " ").title(), "Value": value} for key, value in metrics.items()]
    )
    issues_df = pd.DataFrame(validation_report.get("issues", []))
    return metric_df, issues_df


def _selected_keys(output_keys: Iterable[str] | None, defaults: set[str]) -> set[str]:
    return set(output_keys) if output_keys is not None else set(defaults)


def _excel_writer(path: Path) -> pd.ExcelWriter:
    engine = "xlsxwriter" if importlib.util.find_spec("xlsxwriter") else "openpyxl"
    return pd.ExcelWriter(path, engine=engine)


def _is_xlsxwriter(writer: pd.ExcelWriter) -> bool:
    return getattr(writer, "engine", "") == "xlsxwriter"


def _hex_color(color: str) -> str:
    color = color.strip()
    return color if color.startswith("#") else f"#{color}"


def _xlsxwriter_formats(writer: pd.ExcelWriter) -> dict[str, object]:
    cache = getattr(writer, "_workbook_formats", None)
    if cache is None:
        workbook = writer.book
        cache = {
            "header": workbook.add_format(
                {
                    "bold": True,
                    "font_color": _hex_color(HEADER_FONT),
                    "bg_color": _hex_color(HEADER_FILL),
                    "align": "center",
                    "valign": "vcenter",
                }
            ),
            "summary_light": workbook.add_format({"bg_color": _hex_color(LIGHT_FILL)}),
        }
        setattr(writer, "_workbook_formats", cache)
    return cache


def _column_widths_from_frame(df: pd.DataFrame) -> list[int]:
    sample = df.head(WIDTH_SCAN_ROW_LIMIT)
    widths: list[int] = []
    for column_index, column in enumerate(df.columns):
        max_length = len(str(column or ""))
        if not sample.empty:
            values = sample.iloc[:, column_index].fillna("").astype(str)
            if not values.empty:
                max_length = max(max_length, int(values.map(len).max()))
        widths.append(min(max(max_length + 2, 12), 42))
    return widths


def _style_xlsxwriter_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    worksheet = writer.sheets[sheet_name]
    formats = _xlsxwriter_formats(writer)
    worksheet.freeze_panes(1, 0)
    worksheet.hide_gridlines(2)
    row_count, column_count = df.shape
    if column_count:
        worksheet.autofilter(0, 0, row_count, column_count - 1)
    for column_index, column_name in enumerate(df.columns):
        worksheet.write(0, column_index, column_name, formats["header"])
    for column_index, width in enumerate(_column_widths_from_frame(df)):
        worksheet.set_column(column_index, column_index, width)
    if sheet_name == "Summary":
        for row_index in range(1, row_count + 1, 2):
            worksheet.set_row(row_index, None, formats["summary_light"])


def _write_dataframe(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    if _is_xlsxwriter(writer):
        _style_xlsxwriter_sheet(writer, sheet_name, df)


def _enhance_xlsxwriter_district_workbook(
    writer: pd.ExcelWriter,
    summary_tables: dict[str, pd.DataFrame],
    district_sheets: list[tuple[str, pd.DataFrame]] | None = None,
    cluster_col: str | None = None,
) -> None:
    workbook = writer.book
    summary_df = summary_tables["district_summary"]
    if "Summary" in writer.sheets and len(summary_df) and summary_df.shape[1] >= 5:
        worksheet = writer.sheets["Summary"]
        chart = workbook.add_chart({"type": "bar"})
        last_row = len(summary_df)
        chart.add_series(
            {
                "name": ["Summary", 0, 4],
                "categories": ["Summary", 1, 0, last_row, 0],
                "values": ["Summary", 1, 4, last_row, 4],
            }
        )
        chart.set_title({"name": "Beneficiaries by District"})
        chart.set_y_axis({"name": "District"})
        chart.set_x_axis({"name": "Beneficiaries"})
        chart.set_size({"width": 720, "height": 320})
        worksheet.insert_chart("H2", chart)

    # Shared across the Cluster Summary sheet and every per-district sheet
    # below so xlsxwriter reuses one format per color instead of creating a
    # new one per row (xlsxwriter has a hard cap on formats per workbook).
    color_formats: dict[str, object] = {}

    cluster_df = summary_tables["cluster_summary"]
    if "Cluster Summary" in writer.sheets and "Cluster" in cluster_df.columns:
        worksheet = writer.sheets["Cluster Summary"]
        for row_offset, cluster in enumerate(cluster_df["Cluster"].fillna("").astype(str), start=1):
            color = _cluster_color(cluster)
            if not color:
                continue
            row_format = color_formats.setdefault(
                color, workbook.add_format({"bg_color": _hex_color(color), "font_color": "#000000"})
            )
            worksheet.set_row(row_offset, None, row_format)

    # Color each response row within its own district sheet by cluster too,
    # not just the aggregate Cluster Summary sheet - xlsxwriter can't read
    # back cells it already wrote, so this needs the source dataframe rather
    # than re-reading the sheet the way the openpyxl path below does.
    if district_sheets and cluster_col:
        for sheet_name, group in district_sheets:
            if cluster_col not in group.columns or sheet_name not in writer.sheets:
                continue
            worksheet = writer.sheets[sheet_name]
            for row_offset, cluster in enumerate(group[cluster_col].fillna("").astype(str), start=1):
                color = _cluster_color(cluster)
                if not color:
                    continue
                row_format = color_formats.setdefault(
                    color, workbook.add_format({"bg_color": _hex_color(color), "font_color": "#000000"})
                )
                worksheet.set_row(row_offset, None, row_format)


def _set_column_widths(worksheet) -> None:
    max_row = min(worksheet.max_row, WIDTH_SCAN_ROW_LIMIT)
    for column_index in range(1, worksheet.max_column + 1):
        max_length = 0
        for (value,) in worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            max_length = max(max_length, len(str(value or "")))
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _style_workbook(workbook_or_path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    should_save = isinstance(workbook_or_path, (str, Path))
    path = Path(workbook_or_path) if should_save else None
    workbook = load_workbook(path) if should_save and path else workbook_or_path
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(color=HEADER_FONT, bold=True)
    light_fill = PatternFill("solid", fgColor=LIGHT_FILL)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        _set_column_widths(worksheet)
        if worksheet.title == "Summary":
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                if row[0].row % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill

    if should_save and path:
        workbook.save(path)


def export_cleaned_excel(
    processed_df: pd.DataFrame,
    matches_df: pd.DataFrame | None = None,
    validation_report: dict[str, object] | None = None,
) -> Path:
    path = output_path("settlement_cleaned_response.xlsx")
    metric_df, issues_df = _validation_frames(validation_report)
    with _excel_writer(path) as writer:
        _write_dataframe(writer, processed_df, "Cleaned Response")
        if matches_df is not None and not matches_df.empty:
            _write_dataframe(writer, matches_df, "Settlement Matches")
        if not metric_df.empty:
            _write_dataframe(writer, metric_df, "Validation Metrics")
        if not issues_df.empty:
            _write_dataframe(writer, issues_df, "Validation Issues")
        if not _is_xlsxwriter(writer):
            _style_workbook(writer.book)
    return path


def export_district_workbook(processed_df: pd.DataFrame) -> Path:
    path = output_path("settlement_district_response_workbook.xlsx")
    summary_tables = build_summary_tables(processed_df)
    _, groups = district_groups(processed_df)
    used_sheet_names: set[str] = set()
    cluster_col = detect_column_map(processed_df).get("cluster")

    with _excel_writer(path) as writer:
        _write_dataframe(writer, summary_tables["district_summary"], "Summary")
        _write_dataframe(writer, summary_tables["cluster_summary"], "Cluster Summary")
        _write_dataframe(writer, summary_tables["partner_summary"], "Partner Summary")

        district_sheets: list[tuple[str, pd.DataFrame]] = []
        for district, group in groups:
            sheet_name = truncate_sheet_name(district, used_sheet_names)
            _write_dataframe(writer, group, sheet_name)
            district_sheets.append((sheet_name, group))

        if _is_xlsxwriter(writer):
            _enhance_xlsxwriter_district_workbook(writer, summary_tables, district_sheets, cluster_col)
        else:
            _enhance_district_workbook(writer.book, [name for name, _ in district_sheets], cluster_col)
    return path


def export_district_summary(processed_df: pd.DataFrame) -> Path:
    path = output_path("settlement_district_summary.xlsx")
    summary_tables = build_summary_tables(processed_df)
    with _excel_writer(path) as writer:
        _write_dataframe(writer, summary_tables["district_summary"], "District Summary")
        _write_dataframe(writer, summary_tables["cluster_summary"], "Cluster Summary")
        _write_dataframe(writer, summary_tables["partner_summary"], "Partner Summary")
        if not _is_xlsxwriter(writer):
            _style_workbook(writer.book)
    return path


def _color_sheet_rows_by_cluster(ws, header_name: str) -> None:
    from openpyxl.styles import Font, PatternFill

    cluster_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip().lower() == header_name.strip().lower():
            cluster_col = idx
            break
    if not cluster_col:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cluster = str(row[cluster_col - 1].value or "")
        fill_color = _cluster_color(cluster)
        if fill_color:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.font = Font(color="000000")


def _enhance_district_workbook(
    workbook_or_path,
    district_sheet_names: list[str] | None = None,
    cluster_col: str | None = None,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference

    should_save = isinstance(workbook_or_path, (str, Path))
    path = Path(workbook_or_path) if should_save else None
    workbook = load_workbook(path) if should_save and path else workbook_or_path
    _style_workbook(workbook)

    if "Summary" in workbook.sheetnames:
        ws = workbook["Summary"]
        if ws.max_row > 1 and ws.max_column >= 5:
            chart = BarChart()
            chart.title = "Beneficiaries by District"
            chart.y_axis.title = "Beneficiaries"
            chart.x_axis.title = "District"
            data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "H2")

    if "Cluster Summary" in workbook.sheetnames:
        _color_sheet_rows_by_cluster(workbook["Cluster Summary"], "Cluster")

    # Color each response row within its own district sheet by cluster too,
    # not just the aggregate Cluster Summary sheet.
    if district_sheet_names and cluster_col:
        for sheet_name in district_sheet_names:
            if sheet_name in workbook.sheetnames:
                _color_sheet_rows_by_cluster(workbook[sheet_name], cluster_col)

    if should_save and path:
        workbook.save(path)


def export_all_excel_outputs(
    processed_df: pd.DataFrame,
    matches_df: pd.DataFrame | None,
    validation_report: dict[str, object] | None,
    output_keys: Iterable[str] | None = None,
) -> dict[str, str]:
    selected = _selected_keys(output_keys, EXCEL_OUTPUT_KEYS)
    outputs: dict[str, str] = {}
    if "Cleaned Excel" in selected:
        outputs["Cleaned Excel"] = str(export_cleaned_excel(processed_df, matches_df, validation_report))
    if "District Workbook" in selected:
        outputs["District Workbook"] = str(export_district_workbook(processed_df))
    if "District Summary" in selected:
        outputs["District Summary"] = str(export_district_summary(processed_df))
    return outputs
